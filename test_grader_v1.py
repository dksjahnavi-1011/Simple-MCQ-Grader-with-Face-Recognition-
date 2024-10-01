#Finalised one
import numpy as np
import cv2
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from PIL import Image, ImageTk
from imutils.perspective import four_point_transform
from imutils import contours
import imutils
import os
import openpyxl
import pickle

class ExamGraderApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Automated Grading System")
        self.master.configure(bg="#f0f0f0")

        self.result_text = tk.StringVar()
        self.face_recognizer = FaceRecognizer()

        self.main_frame = tk.Frame(self.master, bg="#f0f0f0", width=600, height=400)
        self.main_frame.pack(expand=True)

        self.title_label = tk.Label(self.main_frame, text="AUTOMATED GRADING SYSTEM!", font=("Arial", 22, "bold"), bg="#f0f0f0")
        self.title_label.grid(row=0, column=0, columnspan=2, pady=20)

        self.panel = tk.Label(self.main_frame, bg="#f0f0f0")
        self.panel.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

        self.btn_select = ttk.Button(self.main_frame, text="Select Image", command=self.process_image, style="Custom.TButton", width=15)
        self.btn_select.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        self.btn_check_grades = ttk.Button(self.main_frame, text="Check Grades", command=self.check_grades, style="Custom.TButton", width=15)
        self.btn_check_grades.grid(row=4, column=0, columnspan=2, padx=5, pady=5)

        self.btn_clear = ttk.Button(self.main_frame, text="Clear", command=self.clear_interface, style="Custom.TButton", width=15)
        self.btn_clear.grid(row=6, column=0, columnspan=2, padx=5, pady=5)

        self.lbl_result = tk.Label(self.main_frame, textvariable=self.result_text, font=("Arial", 14, "bold"), bg="#f0f0f0", fg="green")
        self.lbl_result.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

        self.lbl_score_display = tk.Label(self.main_frame, text="", font=("Arial", 14,"bold"), bg="#f0f0f0", fg="green")
        self.lbl_score_display.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        self.style = ttk.Style()
        self.style.configure("Custom.TButton", foreground="blue", background="#b3d9ff", font=("Arial", 12, "bold"))

    def clear_interface(self):
        # Clearing the result text
        self.result_text.set("")

        # Clearing the score display label
        self.lbl_score_display.config(text="")

        # Clearing the panel
        self.panel.config(image="")

    def process_image(self):
        path = filedialog.askopenfilename()
        if path:
            filename = os.path.basename(path)
            student_name, _ = os.path.splitext(filename)

            image = cv2.imread(path)
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)
            edged = cv2.Canny(blurred, 75, 200)

            cnts = cv2.findContours(edged.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = imutils.grab_contours(cnts)
            docCnt = None

            if len(cnts) > 0:
                cnts = sorted(cnts, key=cv2.contourArea, reverse=True)
                for c in cnts:
                    peri = cv2.arcLength(c, True)
                    approx = cv2.approxPolyDP(c, 0.02 * peri, True)
                    if len(approx) == 4:
                        docCnt = approx
                        break

            paper = four_point_transform(image, docCnt.reshape(4, 2))
            warped = four_point_transform(gray, docCnt.reshape(4, 2))

            thresh = cv2.threshold(warped, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

            cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = imutils.grab_contours(cnts)
            questionCnts = []

            for c in cnts:
                (x, y, w, h) = cv2.boundingRect(c)
                ar = w / float(h)
                if w >= 20 and h >= 20 and ar >= 0.9 and ar <= 1.1:
                    questionCnts.append(c)

            questionCnts = contours.sort_contours(questionCnts, method="top-to-bottom")[0]
            correct = 0

            for (q, i) in enumerate(np.arange(0, len(questionCnts), 5)):
                cnts = contours.sort_contours(questionCnts[i:i + 5])[0]
                bubbled = None

                for (j, c) in enumerate(cnts):
                    mask = np.zeros(thresh.shape, dtype="uint8")
                    cv2.drawContours(mask, [c], -1, 255, -1)
                    mask = cv2.bitwise_and(thresh, thresh, mask=mask)
                    total = cv2.countNonZero(mask)
                    if bubbled is None or total > bubbled[0]:
                        bubbled = (total, j)

                color = (0, 0, 255)
                k = ANSWER_KEY[q]

                if k == bubbled[1]:
                    color = (0, 255, 0)
                    correct += 1

                cv2.drawContours(paper, [cnts[k]], -1, color, 3)

            score = (correct / 5.0) * 100

            self.result_text.set(f"Student Name: {student_name}\nScore: {score:.2f}%")

            image_rgb = cv2.cvtColor(paper, cv2.COLOR_BGR2RGB)
            image_pil = Image.fromarray(image_rgb)
            image_tk = ImageTk.PhotoImage(image_pil)
            self.panel.configure(image=image_tk)
            self.panel.image = image_tk

            write_to_excel(student_name, score)

    def check_grades(self):
        path = filedialog.askopenfilename()
        if path:
            recognized_name, image_path, confidence = self.face_recognizer.recognize_faces(path)

            if recognized_name:
                score = self.fetch_scores_from_excel(recognized_name)
                if score is not None:
                    confidence_text = f"Confidence: {confidence:.2f}%"
                    self.lbl_score_display.config(text=f"Name: {recognized_name}\nScore: {score:.2f}%", fg="green")

                    if image_path:
                        image_pil = Image.open(image_path)
                        image_tk = ImageTk.PhotoImage(image_pil)
                        self.panel.configure(image=image_tk)
                        self.panel.image = image_tk

                        image = cv2.imread(image_path)

                        text = "{} - {:.2f}%".format(recognized_name, score)
                        y = image.shape[0] - 10 if image.shape[0] - 10 > 10 else image.shape[0] + 10

                        face_box = self.face_recognizer.detect_face(image_path)
                        if face_box is not None:
                            (startX, startY, endX, endY) = face_box
                            cv2.rectangle(image, (startX, startY), (endX, endY), (0, 255, 0), 2)
                            y -= 25

                        cv2.rectangle(image, (0, image.shape[0]), (image.shape[1], y), (0, 0, 255), cv2.FILLED)
                        cv2.putText(image, text, (10, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 0), 2)
                        
                        # Adding confidence text to the image
                        cv2.putText(image, confidence_text, (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 255), 2)

                        cv2.imshow("Identified Face", image)
                        cv2.waitKey(3000)
                        cv2.destroyWindow("Identified Face")
                else:
                    self.lbl_score_display.config(text=f"No score found for {recognized_name}.", fg="red")
            else:
                self.lbl_score_display.config(text="No face recognized", fg="red")

    def fetch_scores_from_excel(self, student_name):
        workbook_path = "grades.xlsx"
        if os.path.exists(workbook_path):
            workbook = openpyxl.load_workbook(workbook_path)
            worksheet = workbook.active

            for row in worksheet.iter_rows(values_only=True):
                if row[0] == student_name:
                    workbook.close()
                    return row[1]
            workbook.close()

        return None

class FaceRecognizer:
    def __init__(self):
        detector_path = "face_detection_model"
        embedding_model_path = "nn4.small2.v1.t7"
        recognizer_model_path = "output/recognizer.pickle"
        label_encoder_path = "output/le.pickle"
        self.confidence = 0.5
        
        self.detector = cv2.dnn.readNetFromCaffe(os.path.join(detector_path, "deploy.prototxt"),
                                                  os.path.join(detector_path, "res10_300x300_ssd_iter_140000.caffemodel"))
        self.embedder = cv2.dnn.readNetFromTorch(embedding_model_path)
        self.recognizer = pickle.loads(open(recognizer_model_path, "rb").read())
        self.le = pickle.loads(open(label_encoder_path, "rb").read())

    def recognize_faces(self, image_path):
        image = cv2.imread(image_path)
        image = imutils.resize(image, width=600)
        (h, w) = image.shape[:2]

        image_blob = cv2.dnn.blobFromImage(
            cv2.resize(image, (300, 300)), 1.0, (300, 300),
            (104.0, 177.0, 123.0), swapRB=False, crop=False)

        self.detector.setInput(image_blob)
        detections = self.detector.forward()

        recognized_name = None
        identified_image_path = None
        max_confidence = 0

        for i in range(0, detections.shape[2]):
            confidence = detections[0, 0, i, 2]

            if confidence > self.confidence and confidence > max_confidence:
                max_confidence = confidence
                box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
                (startX, startY, endX, endY) = box.astype("int")

                face = image[startY:endY, startX:endX]
                (fH, fW) = face.shape[:2]

                if fW < 20 or fH < 20:
                    continue

                face_blob = cv2.dnn.blobFromImage(face, 1.0 / 255, (96, 96), (0, 0, 0), swapRB=True, crop=False)
                self.embedder.setInput(face_blob)
                vec = self.embedder.forward()

                preds = self.recognizer.predict_proba(vec)[0]
                j = np.argmax(preds)
                proba = preds[j]
                name = self.le.classes_[j]

                recognized_name = name
                identified_image_path = image_path

        return recognized_name, identified_image_path, max_confidence*100

    def detect_face(self, image_path):
        image = cv2.imread(image_path)
        (h, w) = image.shape[:2]

        image_blob = cv2.dnn.blobFromImage(
            cv2.resize(image, (300, 300)), 1.0, (300, 300),
            (104.0, 177.0, 123.0), swapRB=False, crop=False)

        self.detector.setInput(image_blob)
        detections = self.detector.forward()

        face_box = None

        for i in range(0, detections.shape[2]):
            confidence = detections[0, 0, i, 2]

            if confidence > self.confidence:
                box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
                face_box = box.astype("int")
                break

        return face_box

def write_to_excel(student_name, score):
    workbook_path = "grades.xlsx"
    if os.path.exists(workbook_path):
        workbook = openpyxl.load_workbook(workbook_path)
    else:
        workbook = openpyxl.Workbook()

    worksheet = workbook.active
    
    # Check if student_name already exists in the Excel sheet
    existing_names = [cell.value for cell in worksheet['A']]
    if student_name in existing_names:
        messagebox.showerror("Duplicate Entry", f"{student_name} already exists in the Excel sheet.")
        workbook.close()
        return

    # If student_name doesn't exist, append the new entry
    worksheet.append([student_name, score])
    workbook.save(workbook_path)
    

ANSWER_KEY = {0: 1, 1: 4, 2: 0, 3: 3, 4: 1}

def main():
    root = tk.Tk()
    app = ExamGraderApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
